import os
import shutil
from openpyxl import load_workbook
from utils import isDirectory, createDirectory


def readWorkbook(fileName):
    """

    :param fileName:
    :return:
    """
    return load_workbook(filename=fileName)


def printWorkbook(fileName, dstDir):
    """

    :param fileName:
    :return:
    """
    if not isDirectory(dstDir):
        createDirectory(dstDir)

    os.startfile(fileName, 'print')


def printWorksheets(srcFileName, dstDirName):
    """

    :param srcFileName:
    :param dstDirName:
    :return:
    """
    if not isDirectory(dstDirName):
        createDirectory(dstDirName)

    wb = readWorkbook(srcFileName)

    for sheet in wb:
        dstFileName = dstDirName + f'\\{sheet.title}.xlsx'
        print(dstFileName)

        shutil.copyfile(srcFileName, dstFileName)
        _wb = readWorkbook(dstFileName)

        for _sheet in _wb:
            if _sheet.title != sheet.title:
                _wb.remove(_sheet)
        _wb.save(dstFileName)

        printWorkbook(dstFileName, dstDirName)


if __name__ == '__main__':
    srcFileName = 'C:\\Users\\cmolina\\Downloads\\test spreadsheet.xlsx'
    dstDirName = 'C:\\Users\\cmolina\\Downloads\\submittal\\cover sheets'

    printWorksheets(srcFileName, dstDirName)
